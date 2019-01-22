'''
====================================================================
Library for reporting a test sequence and results in html format
(C) Copyright 2018 Lear Corporation
====================================================================
'''
__author__  = 'Pradeep Singh'
__version__ = '1.0.0'
__email__   = 'psingh02@lear.com'

'''
CHANGE LOG
==========
1.0.0 Initial version.
'''

import HTML
import datetime
import fileinput
import os
from openpyxl import *




class HTMLReport(object):

    def __init__(self, tlaName='', projName='', swVersion='', hwVersion='',network='', author=''):
        '''
        Description: Constructor.
        Parameter 'tlaName' is a string with the name of the SW being tested.
        Parameter 'swVersion' is optional, string with the SW branch, in case it's useful to identify it.
        Parameter 'hwVersion' is optional, string with the HW version used to do the test in the target.
        Parameter 'author' is optional, string with author of the test script.
        '''
        # Initialize headers and other info for the HTML report
        self.__initialize_html()
        self.all_testcases = []
        self.all_teststeps = []
        self.all_passfail  = []
        self.tt_oks        = 0
        self.tt_noks       = 0
        # Store parameters
        self.tla_name      = tlaName
        self.prj_name      = projName
        self.sw_ver        = swVersion
        self.hw_ver        = hwVersion
        self.network       = network
        self.author        = author


    def __initialize_html(self):
        '''
        Description: Initializes some tags and table properties for html file.
                     Must be used within the class.
                     Not accessible outside class as it's a private function.

        '''
        #------------------------------FOR SUMMARY REPORT-------------------------------------------#
        self.summaryHtmlHeaderStr = \
        """<DIV STYLE="background-color:#0080C0; color:white; padding:4px;font-size:25pt">
                <SPAN STYLE="font-weight:bold; color:white;font-size:15pt">
                    <CENTER>AUTOMATED ECU TEST REPORT</CENTER>
                </SPAN>
        </DIV>"""

        self.summaryHtmlInfoBlockStr = \
        """
        <DIV style="background-color:#ffffff;height:auto" >
        <DIV class="configblock" style="display:inline-block;position:relative;float:left;"><b>INFORMATION</b> </DIV>
        <DIV style="width:30pt;display:inline-block;position:relative;"></DIV>
        """

        self.summaryHtmlTestBlockStr = \
        """
        <DIV style="background-color:#ffffff;height:auto" >
        <DIV class="configblock" style="display:inline-block;position:relative;float:left;"><b>TEST SUMMARY</b> </DIV>
        <DIV style="width:20pt;display:inline-block;position:relative;"></DIV>
        """


        self.summaryInformationData    = [['Project :',''],['Date :',''],['Time :',''],['SW Version :',''],['HW Version :',''],['Network :',''],['Author :','']]
        self.summaryTestData           = [['No. of Tests :',''],['Tests Passed :',''],['Tests Failed :','']]
        self.summaryTestListColHeaders = ['Sr No', 'List Of Tests', 'Total Tests', 'Tests Passed', 'Tests Failed']
        self.summaryTestListColWidth   = [64, 440, 100, 100, 100]
        self.summaryTestListColStyles  = ['background-color:white','background-color:white','background-color:white','background-color:white','background-color:white']
        self.summaryTestListColAlign   = ['center','left','center','center','center']
        self.summaryTestListBuffer     = []

        #------------------------------FOR INDIVIDUAL TEST REPORT-------------------------------------------#
        self.htmlHeaderStr = \
        """
        <DIV STYLE="background-color:#ffffff; color:white; padding:4px;font-size:25pt">
                <SPAN STYLE="font-weight:bold; color:black;font-size:20pt">
                    <CENTER>%s</CENTER>
                </SPAN>
        </DIV>
        """

        self.htmlGapStr = """\n<DIV STYLE="background-color:#ffffff;height:20"></DIV>\n"""
        self.htmlHrefStrPre1 = '''<a href="#Test'''
        self.htmlHrefStrPre2 = '''">'''
        self.htmlHrefStrPost = '''</a>'''

        self.reportInformationTableData      = [['Project :', ''],['Date :', ''],['Time :', ''],['SW Version :', ''],['HW Version :', ''],['Network :', ''],['Author :', '']]
        self.reportTestListTableHeaders      = ['Sr No', 'Test Cases', 'Test Steps Passed', 'Test Steps Failed']
        self.reportTestCaseTableHeaders      = ['Test Steps', 'Result', 'Actual', 'Expected', 'Comments']
        self.reportInformationTableColWidth  = [200,200]
        self.reportTestListColWidth          = [64, 440, 100, 100]
        self.reportTestCaseColWidth          = [500, 64, 300, 300, 200]
        self.reportInformationTableColStyles = ['font-weight:bold','font-weight:normal']
        self.reportTestListColHeaderStyles   = ['background-color:#E3E3E3','background-color:#E3E3E3','background-color:#24FE2B','background-color:#F53405']
        self.reportTestListColStyles         = ['background-color:white','background-color:white;white-space: pre','background-color:white','background-color:white']
        self.reportTestListColAlign          = ['center','left','center','center']
        self.reportTestCaseColAlign          = ['left','center','left','left','left']
        self.reportTestCaseColStyles         = ['background-color:white;white-space: normal','background-color:white','background-color:white;white-space: pre','background-color:white;white-space: pre','background-color:white;white-space: pre']
        self.reportTestListBuffer            = []


    def __createReportDir(self,directory):
        '''
        Description: Creates Report folder structure.
                     Must be used within the class.
        '''
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print ('Error: Creating directory. ' +  directory)


    def __start_html(self):
        '''
        Description: Creates Report folder structure and Updates the html report
                     file with the test report data.Must be used within the class.
                     Not accessible outside class as it's a private function.
        '''
        self.currentDT    = datetime.datetime.now()
        self.reportDir    = (self.prj_name + '_' + str(self.currentDT.day)+'.'+str(self.currentDT.month)+'.'+str(self.currentDT.year) + '_' + str(self.currentDT.hour)+'.'+str(self.currentDT.minute))
        self.reportSubDir = self.tla_name

        self.__createReportDir('./' + self.reportDir + '/' + self.reportSubDir + '/')

        self.reportFileName = self.tla_name + "_Report.html"
        self.fileObj        = file(('./' + self.reportDir + '/' + self.reportSubDir + '/' + self.reportFileName),'w+')



    def __update_html(self):
        '''
        Description: Updates the html file with the test report data.
                     Must be used within the class.
                     Not accessible outside class as it's a private function.

        '''

        self.reportHeading = self.tla_name + " Test Report"
        self.fileObj.write(self.htmlHeaderStr %self.reportHeading)
        self.fileObj.write(self.htmlGapStr)

        #------------------UPDATE REPORT INFORMATION DATA-------------------------------------#
        self.currentDate = str(self.currentDT.day)+'-'+str(self.currentDT.month)+'-'+str(self.currentDT.year)
        self.currentTime = str(self.currentDT.hour)+':'+str(self.currentDT.minute)

        self.reportInformationTableData[0][1] = self.prj_name
        self.reportInformationTableData[1][1] = self.currentDate
        self.reportInformationTableData[2][1] = self.currentTime
        self.reportInformationTableData[3][1] = self.sw_ver
        self.reportInformationTableData[4][1] = self.hw_ver
        self.reportInformationTableData[5][1] = self.network
        self.reportInformationTableData[6][1] = self.author


        self.fileObj.write('''\n<DIV style="background-color:#ffffff;height:auto" >''')
        self.fileObj.write("\n<CENTER>")
        self.fileObj.write(HTML.table(self.reportInformationTableData, col_width=self.reportInformationTableColWidth, col_styles=self.reportInformationTableColStyles, style="display:inline-block;position:relative;align:center;background-color:#E3E3E3;", border='0'))
        self.fileObj.write("\n</CENTER>")
        self.fileObj.write("\n</DIV>")
        self.fileObj.write(self.htmlGapStr)
        #------------------UPDATE TEST SUMMARY TABLE DATA-------------------------------------#
        self.fileObj.write("\n<DIV><b><u>TEST SUMMARY</u></b></DIV>\n")
        self.fileObj.write(self.htmlGapStr)

        for i in range(0,len(self.all_testcases)):
            testCaseName = self.htmlHrefStrPre1 + str(i+1) + self.htmlHrefStrPre2 + self.all_testcases[i][0] + self.htmlHrefStrPost
            self.reportTestListBuffer.append([str(i+1),testCaseName, self.all_passfail[i][len(self.all_passfail[i])-1][0], self.all_passfail[i][len(self.all_passfail[i])-1][1]])

        self.fileObj.write(HTML.table(self.reportTestListBuffer, header_row=self.reportTestListTableHeaders,col_header_styles=self.reportTestListColHeaderStyles, col_align=self.reportTestListColAlign, col_styles=self.reportTestListColStyles, col_width=self.reportTestListColWidth, cellpadding=0, cellspacing=0, width=1235, style='border-collapse: collapse;table-layout:fixed;width:745pt', border='1'))
        self.fileObj.write(self.htmlGapStr*4)

        #------------------UPDATE TEST CASES TABLE DATA-------------------------------------#
        self.fileObj.write("\n<DIV><b><u>TEST CASES</u></b></DIV>\n")
        self.fileObj.write(self.htmlGapStr*2)

        for idx in range(0,len(self.all_testcases)):
            self.fileObj.write('''\n<DIV style="white-space: pre" id="Test%s"><b>%s</b></DIV>''' %(str(idx+1),self.all_testcases[idx][0]))
            self.fileObj.write('''\n<DIV style="white-space: pre">%s</DIV>''' %self.all_testcases[idx][1])
            self.fileObj.write("\n")
            self.fileObj.write(HTML.table(self.all_teststeps[idx], header_row=self.reportTestCaseTableHeaders, col_width=self.reportTestCaseColWidth, col_align=self.reportTestCaseColAlign, col_styles=self.reportTestCaseColStyles, cellpadding=0, cellspacing=0, width=1235, style='border-collapse: collapse;table-layout:fixed;width:1100pt;background-color:#E3E3E3', border='1'))
            self.fileObj.write(self.htmlGapStr*4)


    def __finish_html(self):
        '''
        Description: Closes the html file.Must be used within the class.
                     Not accessible outside class as it's a private function.
        '''
        self.fileObj.close()

        self.tempFileObj = fileinput.FileInput(('./' + self.reportDir + '/' + self.reportSubDir + '/' + self.reportFileName), inplace=1)
        for line in self.tempFileObj:
            if ">OK<" in line:
                line = line.replace('background-color:white','background-color:#24FE2B')
            elif ">NOK<" in line:
                line = line.replace('background-color:white','background-color:#F53405')

            line = line.rstrip('\r\n')

            print line

        self.tempFileObj.close()
        self.__generate_summary_report()



    def __generate_summary_report(self):
        '''
        Description: Generates the html file with the test report data.
                     Must be called at the end of the script.
        '''
        self.summaryReport  = "Summary.html"
        self.summaryFileObj = file('./' + self.reportDir +'/'+ self.summaryReport,'w+')

        #-------------------------SUMMARY REPORT HEADING-------------------------------------#
        self.summaryFileObj.write(self.summaryHtmlHeaderStr)
        self.summaryFileObj.write(self.htmlGapStr)

        #-------------------------UPDATE SUMMARY REPORT INFORMATION DATA----------------------#
        self.summaryInformationData[0][1]     = self.prj_name
        self.summaryInformationData[1][1]     = self.currentDate
        self.summaryInformationData[2][1]     = self.currentTime
        self.summaryInformationData[3][1]     = self.sw_ver
        self.summaryInformationData[4][1]     = self.hw_ver
        self.summaryInformationData[5][1]     = self.network
        self.summaryInformationData[6][1]     = self.author

        self.summaryFileObj.write(self.summaryHtmlInfoBlockStr)
        self.summaryFileObj.write(HTML.table(self.summaryInformationData,style="display:inline-block;position:relative;background-color:#E3E3E3;",border='0'))
        self.summaryFileObj.write("\n</DIV>")
        self.summaryFileObj.write(self.htmlGapStr)

        #-------------------------UPDATE TEST SUMMARY DATA----------------------------------#
        self.summaryTestData[0][1] = str(self.tt_oks + self.tt_noks)
        self.summaryTestData[1][1] = str(self.tt_oks)
        self.summaryTestData[2][1] = str(self.tt_noks)

        self.summaryFileObj.write(self.summaryHtmlTestBlockStr)
        self.summaryFileObj.write(HTML.table(self.summaryTestData,style="display:inline-block;position:relative;background-color:#E3E3E3;",border='0'))
        self.summaryFileObj.write("\n</DIV>")
        self.summaryFileObj.write(self.htmlGapStr)


        #------------------UPDATE TEST CASES LIST TABLE -------------------------------------#
        hrefLinkPre = ('''<a href="%s"  target="_blank">''' %(self.reportSubDir +'/'+ self.reportFileName))
        hrefLinkPost = '''</a>'''
        self.summaryTestListBuffer.append([1,(hrefLinkPre + self.tla_name + hrefLinkPost), str(self.tt_oks + self.tt_noks), str(self.tt_oks), str(self.tt_noks)])
        self.summaryFileObj.write(HTML.table(self.summaryTestListBuffer, header_row=self.summaryTestListColHeaders,col_align=self.summaryTestListColAlign, col_styles=self.summaryTestListColStyles, col_width=self.summaryTestListColWidth, cellpadding=0, cellspacing=0, width=1235, style='border-collapse: collapse;table-layout:fixed;width:927pt;background-color:#0080C0', border='1'))

        self.summaryFileObj.close()

        #Invoke Summary Report html file
        os.startfile(os.getcwd() + '/' + self.reportDir +'/'+ self.summaryReport)



    def add_test_case(self, title='', description='', reqs=''):
        '''
        Description: Adds a new test case (also called test scenario) to the report.
        Parameter 'title' is a string with the name of the test case.
        Parameter 'description' is optional, a string with a short description of the test case.
        Parameter 'reqs' is optional, a string with the requirements tested in this test case.
        If there are no requirements tested, just leave it as an empty string.

        Example:
            it_report = ITReport('FrontWiper', '1.0.0', 'MY_2015', '5271', 'HW6.1', 'John Doe')
            test_case_name = 'Front Wiper low speed'
            test_case_desc = 'Testing low speed in different key positions'
            test_case_reqs = 'Req_123v2, Req_134'
            it_report.add_test_case(test_case_name, test_case_desc, test_case_reqs)
        '''
        # Initialize headers and other info
        self.tc_oks         = 0
        self.tc_noks        = 0
        # Store parameters
        self.tc_title       = title
        self.tc_description = description
        self.tc_reqs        = reqs

        self.all_testcases.append([self.tc_title, self.tc_description, self.tc_reqs])

        if len(self.all_testcases) > 0:
            self.index = len(self.all_testcases) - 1
        else:
            self.index = 0

        self.all_teststeps.insert(self.index, [])
        self.all_passfail.insert(self.index, [])



    def add_test_step(self, description, result=None, actual='',expected='',comment=''):
        '''
        Description: Adds a new test step in the current test case.
        Parameter 'description' is a string with a short description.
        Parameter 'result' is optional, a boolen indicating the result of the test step.
        A True will generate a OK test, a False will generate a NOK test, and None will generate NOT_TESTED.
        Parameter 'comment' is optional, a string with any interesting comment.

        Note: This method is available for both UTReport and ITReport classes.

        Example:
            it_report = ITReport('FrontWiper', '1.0.0', 'MY_2015', '5271', 'HW6.1', 'John Doe')
            test_case_name = 'Front Wiper low speed'
            test_case_desc = 'Testing low speed in different key positions'
            test_case_reqs = 'Req_123v2, Req_134'
            it_report.add_test_case(test_case_name, test_case_desc, test_case_reqs)
            # Adding a test step
            result = check_act_key_out()
            test_step_description = 'Checking here Front Wiper low speed is not activated in key out'
            it_report.add_test_step(test_step_description, result)
        '''
        # Store parameters
        self.ts_desc     = description
        self.ts_actual   = actual
        self.ts_expected = expected
        self.ts_comment  = comment

        if result == True:
            self.ts_result = 'OK'
            self.tc_oks += 1
            self.tt_oks += 1
        else:
            self.ts_result = 'NOK'
            self.tc_noks += 1
            self.tt_noks += 1

        self.all_teststeps[self.index].append([self.ts_desc, self.ts_result, self.ts_actual, self.ts_expected, self.ts_comment])
        self.all_passfail[self.index].append([self.tc_oks, self.tc_noks])

    def generate_report(self):
        '''
        Description: Generates the html file with the test report data.
                     Must be called at the end of the script.

        Example:
            it_report = ITReport('FrontWiper', '1.0.0', 'MY_2015', '5271', 'HW6.1', 'John Doe')
            test_case_name = 'Front Wiper low speed'
            test_case_desc = 'Testing low speed in different key positions'
            test_case_reqs = 'Req_123v2, Req_134'
            it_report.add_test_case(test_case_name, test_case_desc, test_case_reqs)
            # Adding a test step
            result = check_act_key_out()
            test_step_description = 'Checking here Front Wiper low speed is not activated in key out'
            it_report.add_test_step(test_step_description, result)
            # Generate report
            it_report.generate_report()
        '''
        # Generate stats for the latest test case
#         for i in range(0,len(self.all_testcases)):
#             print self.all_testcases[i]
#             print self.all_teststeps[i]
#             print self.all_passfail[i]

        self.__start_html()
        self.__update_html()
        self.__finish_html()

    def get_log_dir(self):
        return (os.getcwd() + '/' + self.reportDir + '/' + self.reportSubDir + '/' + self.reportSubDir + '_logfile.txt')






class EXCELReport(object):

    def __init__(self, tlaName='', projName='', swVersion='', hwVersion='',network='', author=''):

        self.row_number   = 13 #First Row from where the data insertion will begin in results template
        self.rowTypeCol   = 1
        self.reqIdCol     = 2
        self.testNameCol  = 3
        self.testDescCol  = 4
        self.testCondCol  = 5
        self.expResCol    = 6
        self.actResCol    = 7
        self.testResCol   = 8
        self.commentsCol  = 9

        self.tlaName      = tlaName
        self.prjName      = projName
        self.swVer        = swVersion
        self.hwVer        = hwVersion
        self.networkType  = network
        self.authorName   = author

        self.reportTLA_path   = '../../../Libs/report/reportTLA.xlsx'
        self.reportFileName   = self.tlaName + '_Report.xlsx'
        self.reportWorkbook   = load_workbook(self.reportTLA_path)
        self.reportWorkbook.save(self.reportFileName)

        #Get Worksheet Handle
        self.reportWorksheet = self.reportWorkbook['Sheet1']
        self.reportWorksheet.title = self.tlaName
        self.reportWorkbook.save(self.reportFileName)


    def add_test_case(self, title='', description='', reqs=''):
        '''
        Description: Adds a new test case (also called test scenario) to the report.
        Parameter 'title' is a string with the name of the test case.
        Parameter 'description' is optional, a string with a short description of the test case.
        Parameter 'reqs' is optional, a string with the requirements tested in this test case.
        If there are no requirements tested, just leave it as an empty string.

        Example:
            it_report = EXCELReport('FrontWiper', 'MY_2015', '1.0.0', 'HW1', 'CAN', 'John Doe')
            test_case_name = 'Front Wiper low speed'
            test_case_desc = 'Testing low speed in different key positions'
            test_case_reqs = 'Req_123v2, Req_134'
            it_report.add_test_case(test_case_name, test_case_desc, test_case_reqs)
        '''
        # Initialize headers and other info
        self.tc_oks         = 0
        self.tc_noks        = 0
        # Store parameters
        self.tc_title       = title
        self.tc_description = description
        self.tc_reqs        = reqs

        #Write Testcase data to Worksheet
        self.reportWorksheet.cell(row = self.row_number, column = self.rowTypeCol,  value='TH')
        self.reportWorksheet.cell(row = self.row_number, column = self.reqIdCol,    value=self.tc_reqs)
        self.reportWorksheet.cell(row = self.row_number, column = self.testNameCol, value=self.tc_title)
        self.reportWorksheet.cell(row = self.row_number, column = self.testDescCol, value=self.tc_description)
        #Save Workbook
        self.reportWorkbook.save(self.reportFileName)
        #Increment current rownumber for Worksheet
        self.row_number+=1


    def add_test_step(self, description='', condition='', expected='', actual='', result=None, comment=''):
        '''
        Description: Adds a new test step in the current test case.
        Parameter 'description' is a string with a short description.
        Parameter 'result' is optional, a boolean indicating the result of the test step.
        A True will generate a OK test, a False will generate a NOK test, and None will generate NOT_TESTED.
        Parameter 'comment' is optional, a string with any interesting comment.

        Note: This method is available for both UTReport and ITReport classes.

        Example:
            it_report = EXCELReport('FrontWiper', 'MY_2015', '1.0.0', 'HW1', 'CAN', 'John Doe')
            test_case_name = 'Front Wiper low speed'
            test_case_desc = 'Testing low speed in different key positions'
            test_case_reqs = 'Req_123v2, Req_134'
            it_report.add_test_case(test_case_name, test_case_desc, test_case_reqs)
            # Adding a test step
            result = check_act_key_out()
            test_step_description = 'Checking here Front Wiper low speed is not activated in key out'
            it_report.add_test_step(test_step_description, result)
        '''
        # Store parameters
        self.ts_desc       = description
        self.ts_conditions = condition
        self.ts_actual     = actual
        self.ts_expected   = expected
        self.ts_comment    = comment

        if result == True:
            self.ts_result = 'OK'
            self.tc_oks += 1
        else:
            self.ts_result = 'NOK'
            self.tc_noks += 1

        #Write Teststep data to Worksheet
        self.reportWorksheet.cell(row = self.row_number, column = self.rowTypeCol,  value='TS')
        self.reportWorksheet.cell(row = self.row_number, column = self.testDescCol, value=self.ts_desc)
        self.reportWorksheet.cell(row = self.row_number, column = self.testCondCol, value=self.ts_conditions)
        self.reportWorksheet.cell(row = self.row_number, column = self.expResCol,   value=self.ts_expected)
        self.reportWorksheet.cell(row = self.row_number, column = self.actResCol,   value=self.ts_actual)
        self.reportWorksheet.cell(row = self.row_number, column = self.testResCol,  value=self.ts_result)
        self.reportWorksheet.cell(row = self.row_number, column = self.commentsCol, value=self.ts_comment)
        #Save Workbook
        self.reportWorkbook.save(self.reportFileName)
        #Increment current rownumber for Worksheet
        self.row_number+=1

    def generate_report(self):
        '''
        Description: Generates the html file with the test report data.
                     Must be called at the end of the script.
        '''
        os.startfile(self.reportFileName)

    def get_log_dir(self):
        return (os.getcwd() + self.tlaName + '_logfile.txt')
