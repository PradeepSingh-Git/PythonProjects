'''
====================================================================
Framework for autogenerating Testcases in Python by fetching data from
excel sheet. These scripts then run the tests automatically on ECU.
(C) Copyright 2018 Lear Corporation
====================================================================
'''


__author__ = 'Pradeep Singh'
__version__ = '1.0.1'
__email__   = 'psingh02@lear.com'

'''
CHANGE LOG
==========
1.0.1 Completely redesigning the template and script files
1.0.0 Initial version
'''

import sys
import os
from openpyxl import *


class Testsuite:
    '''
    Class for accessing the Framework functionalities.
    '''

    def __init__(self,workBookName='',workSheetName='',workBookPath='',tableHdrRow=17,rowTypeCol=0,cmdTypeCol=1,reqIDCol=2,testNameCol=3,testDescCol=4,testCondCol=5,expResCol=6,actResCol=7,testResCol=8,commentCol=9,first_TH_row=18):
        '''
        Description: Constructor. Access the physical communication device.

        Example:
            tsuiteObj = Testsuite()
        '''
        self.workBookName  = workBookName.split('.')[0]
        self.workSheetName = workSheetName
        self.workBookPath  = workBookPath

        self.tableHdrRow = tableHdrRow
        self.rowTypeCol  = rowTypeCol
        self.cmdTypeCol  = cmdTypeCol
        self.reqIDCol    = reqIDCol
        self.testNameCol = testNameCol
        self.testDescCol = testDescCol
        self.testCondCol = testCondCol
        self.expResCol   = expResCol
        self.actResCol   = actResCol
        self.testResCol  = testResCol
        self.commentCol  = commentCol

        self.curr_row    = first_TH_row
        self.Test_Case_Cnt = 0
        self.Test_Step_Cnt = 0
        self.writerow_number=tableHdrRow
        self.cwd =os.getcwd()



#TODO change the functions name at all instances EXCEL PATH

    def readTestWorkbook(self):
        '''
        Description: This function open Excel sheet in which all test cases are written.
        user shoud keep the testcase excel sheet in Testsuite folder.
        '''
        self.book = load_workbook(filename = str(self.workBookPath + "\\" + self.workBookName + '.xlsm'))
        self.xl_sheet = self.book[self.workSheetName]


    def createDataDictionary(self):

        #-------------Check for TH at first pos
        if((str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value)=='TS')):
            print "\n#######################################"
            print "\n##Please  Begin your tests  with a TH##"
            print "\n#######################################"
            sys.exit()
        #--------------------

        if((str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value)=='END Line. Do Not Remove')):
            print "ERROR ADD SOME DATA TO THE SHEET"
        self.list_of_tests=[]
        self.Test_Case_Cnt=0


        while (str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value)!='END Line. Do Not Remove'): #Loop starts to access the test case/steps untill the END of  Line
            #check for beginning with 'TH' or 'TS'
            self.commandType=[]
            self.testDescription=[]
            self.testConditions=[]
            self.expectedResult=[]
            self.actualResult=[]
            self.testResults=[]
            self.comments=[]
            self.Test_Step_Cnt=0
            self.Test_data ={
                             'Row_TYPE' :'',
                             'TS_Command_TYPE':self.commandType,#Addded
                             'TH_Requirement_ID':'',
                             'TH_Test_NAME':'',
                             'TH_Test_DESCRIPTION':'',
                             'TS_Test_DESCRIPTION':self.testDescription,#
                             'TS_Test_CONDITION':self.testConditions,
                             'TS_Expected_RESULT':self.expectedResult,
                             'TS_Actual_RESULT':self.actualResult,
                             'TS_Test_RESULT':self.testResults,
                             'TS_Comments':self.comments,
                             'TH_Tscount':''
                          }
            if(
                     str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) != ('TH')
                and  str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) != ('TS')
               ):

                print "\nWrong data entered in row" + str(self.curr_row) + " column " +str(self.rowTypeCol)+"\n"
                print "The Cells in this column should contain TH or TS only"
                print "Make sure to begin with a TH"
                sys.exit()

                '''for test Cases'''
            if (
                str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == 'TH'
                ):

                print "Found TH/TS in column 1 Ready to Go"
                print("\n*************Filling the data for test case Number")+str(self.Test_Case_Cnt) +("*************")
                '''Row type is required for both'''
                self.Test_Case_Cnt += 1 # Increment the Test Case Count
                self.Test_data['Row_TYPE']              = self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value
                self.Test_data['TH_Requirement_ID']     =str(self.xl_sheet.cell(self.curr_row,self.reqIDCol).value)
                self.Test_data['TH_Test_NAME']          =str(self.xl_sheet.cell(self.curr_row,self.testNameCol).value)
                self.Test_data['TH_Test_DESCRIPTION']   = str(self.xl_sheet.cell(self.curr_row,self.testDescCol).value)#conditions
                self.curr_row += 1
                if str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == 'TS':
                    while(
                             str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == ('TS')
                          or str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == ('')
                          ):
                        if(str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == ('')):
                            self.curr_row +=1
                        elif(str(self.xl_sheet.cell(self.curr_row,self.rowTypeCol).value) == ('TS')):

                            ''''for test Steps'''
                            self.commandType.append(str(self.xl_sheet.cell(self.curr_row,self.cmdTypeCol).value))
                            self.testDescription.append(str(self.xl_sheet.cell(self.curr_row,self.testDescCol).value))#conditions
                            self.testConditions.append(str(self.xl_sheet.cell(self.curr_row,self.testCondCol).value))
                            self.expectedResult.append(str(self.xl_sheet.cell(self.curr_row,self.expResCol).value))
                            self.actualResult.append(str(self.xl_sheet.cell(self.curr_row,self.actResCol).value))
                            self.testResults.append(str(self.xl_sheet.cell(self.curr_row,self.testResCol).value))
                            self.comments.append(str(self.xl_sheet.cell(self.curr_row,self.commentCol).value))
                            print str(self.curr_row)
                            print "Added Test step "+str(self.Test_Step_Cnt)
                            self.Test_Step_Cnt+=1
                            self.curr_row += 1
                    self.Test_data['TH_Tscount']   = str(self.Test_Step_Cnt)
            self.list_of_tests.append(self.Test_data)
            self.totaltests=str(self.Test_Case_Cnt)
            print "Added data of row %s" %self.curr_row


    def get_workbookpath(self):
        return str(self.workBookPath + "\\" + self.workBookName + '.xlsm') #path of excelsheet


    def get_worksheet(self):
        return str(self.workSheetName)


    def get_pyScriptFileName(self):
        return str(self.workBookName + '_' + self.workSheetName + '.py')


    def get_batchScriptFileName(self):
        return str(self.workBookName + '_' + self.workSheetName + '.bat')


    def get_totaltests(self):
        return self.Test_Case_Cnt


    def get_keys(self):
        if(
           self.list_of_tests[0].keys()!=None
           ):
            return self.list_of_tests[0].keys()
        else:
            print "Error\nDictionary not filled\n"


    def get_Title(self):
        return str(self.xl_sheet.cell(4,self.testNameCol).value)


    def get_Author(self):
        return str(self.xl_sheet.cell(6,self.testNameCol).value)


    def get_Project(self):
        return str(self.xl_sheet.cell(8,self.testNameCol).value)


    def get_SoftwareVersion(self):
        return str(self.xl_sheet.cell(10,self.testNameCol).value)


    def get_HardwareVersion(self):
        return str(self.xl_sheet.cell(12,self.testNameCol).value)


    def get_NetworkType(self):
        return str(self.xl_sheet.cell(14,self.testNameCol).value)


    def get_CANChannelID(self):
        return str(self.xl_sheet.cell(4,self.actResCol).value)


    def get_CANChannelBR(self):
        return str(self.xl_sheet.cell(4,self.testResCol).value)


    def get_LINChannelID(self):
        return str(self.xl_sheet.cell(6,self.actResCol).value)


    def get_LINChannelBR(self):
        return str(self.xl_sheet.cell(6,self.testResCol).value)


    def get_FlexrayChannelID(self):
        return str(self.xl_sheet.cell(8,self.actResCol).value)


    def get_FlexrayChannelBR(self):
        return str(self.xl_sheet.cell(8,self.testResCol).value)


    def get_DbcPath(self):
        return str(self.xl_sheet.cell(10,self.actResCol).value)


    def get_testcommandtype(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Command_TYPE'][teststepnumber])


    def get_testrequirementid(self,testcasenumber):
        return str(self.list_of_tests[testcasenumber]['TH_Requirement_ID'])


    def get_testname(self,testcasenumber):
        return str(self.list_of_tests[testcasenumber]['TH_Test_NAME'])


    def get_testdescription(self,testcasenumber):
        return str(self.list_of_tests[testcasenumber]['TH_Test_DESCRIPTION'])


    def get_teststepdescription(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Test_DESCRIPTION'][teststepnumber])


    def get_testconditions(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Test_CONDITION'][teststepnumber])


    def get_expectedresult(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Expected_RESULT'][teststepnumber])


    def get_actualresult(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Actual_RESULT'][teststepnumber])


    def get_testresults(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Test_RESULT'][teststepnumber])


    def get_comments(self,testcasenumber,teststepnumber):
        return str(self.list_of_tests[testcasenumber]['TS_Comments'][teststepnumber])


    def get_numberofteststeps(self,testcasenumber):
        return str(self.list_of_tests[testcasenumber]['TH_Tscount'])

#-------------------------Function Definitons related to Script writing begin from here---------------------------------------------------

#----------------------------Create Folder Structures--------------------------------------------------
    def createFolders(self):
        '''
        Description: This function will create folder structure for each testSheet

        '''
        createDir  = self.workBookPath + "\\" + self.workBookName + "\\" + self.workSheetName

        if not os.path.exists(createDir):
            os.makedirs(createDir)

#----------------------------Opening the python file---------------------------------------------------
    def openPythonFile(self):
        '''
        Description: This function will create new Python file i.e .py file for each testSheet
                     along with corresponding folder structure.
        '''
        pytitle = self.workBookPath + "\\" + self.workBookName + "\\" + self.workSheetName + "\\"  + self.get_pyScriptFileName()
        self.fo = open(pytitle,'w')    #open file in write mode

#-----------------------------Write Batchfile----------------------------------------------------------
    def createBatchFile(self):
        '''
        Description: This Function writes to batch file that executes corresponding .py file
        '''
        batchtitle = self.workBookPath + "\\" + self.workBookName + "\\" + self.workSheetName + "\\"  + self.get_batchScriptFileName()

        self.batchfo = open(batchtitle,'w') #open file in write mode
        self.batchfo.write("C:\Python27\python.exe" + " %s" %(self.get_pyScriptFileName()))
        self.batchfo.close()

#---------------------------- Generate Copyright header -------------------------------------------#
    def writeCopyright(self):
        '''
        Description: This Function import packages in .py file
        '''
        global div
        div = '======================================='
        self.fo.write('#' + div * 4)
        self.fo.write("\n#")
        self.fo.write("\n# Integration Test script file")
        self.fo.write("\n# (c) Copyright 2012 Lear Corporation ")
        self.fo.write("\n#")
        self.fo.write('\n#' + div * 4 + '\n')

        self.fo.write("\nimport sys")
        self.fo.write("\nimport os")
        self.fo.write("\nimport time")
        self.fo.write("\nimport Tkinter as tk")
        self.fo.write("\nimport tkMessageBox")
        self.fo.write("\nfrom openpyxl import *")

#---------------------------- Generate Path Settings -----------------------------------------------#
    def writePathSettings(self):
        '''
        Description: This function write all project related path to py file.
        Example:
             Tools_PATH,CODE_PATH
        creates object for t32,com
        '''
        self.fo.write("\n\nREPORT_API_PATH = os.path.abspath(r'../../../Libs/report')")
        self.fo.write("\nCOM_API_PATH = os.path.abspath(r'../../../Libs/com')")
        self.fo.write("\nUSERDEFINED_API_PATH = os.path.abspath(r'../../../Testsuite/')")

        self.fo.write("\n\n# Adding paths for loading correctly .py libraries")
        self.fo.write("\nsys.path.append(REPORT_API_PATH)")
        self.fo.write("\nsys.path.append(COM_API_PATH)")
        self.fo.write("\nsys.path.append(USERDEFINED_API_PATH)")

        self.fo.write("\n\nfrom report import *")
        self.fo.write("\nfrom com import *")
        self.fo.write("\nfrom user_defined_api import *")

#---------------------------- Generate Report Settings ---------------------------------------------#
    def writeReportSettings(self):
        '''
        Description:This function write all extracted values(from excel sheet) regarding the report header in .py file
        '''
        self.fo.write("\n\n\n# Report Header Variables")
        self.fo.write("\nAUTHOR = '%s'" %self.get_Author())
        self.fo.write("\nTLA = '%s'" % self.get_Title() )
        self.fo.write("\nPROJECT_NAME = '%s'" % self.get_Project() )
        self.fo.write("\nHW_VERSION = '%s'" %  self.get_HardwareVersion())
        self.fo.write("\nNETWORK_TYPE = '%s'" % self.get_NetworkType())
        self.fo.write("\nSW_VERSION = '%s'" % self.get_SoftwareVersion())
        self.fo.write("\n\n# Create the report object, specifying the test data")
        self.fo.write("\nreport = EXCELReport(TLA, PROJECT_NAME, SW_VERSION, HW_VERSION, NETWORK_TYPE, AUTHOR)")

#---------------------------Wrtie Channel setting to the py file------------------------------------
    def writeChannelSettings(self):
        '''
        Description:This for Loop Extract can channel baud rate,Channel ID,LDf,DBC.
        it Checks CAN Access req. and according to that set channel
        '''
        self.fo.write("\n\n")
        self.fo.write("#################################################################\n")
        self.fo.write("##               Set CAN,LIN,FR Channels                       ##\n")
        self.fo.write("#################################################################\n")
        self.fo.write("\ncom = Com('VECTOR')")
        self.fo.write("\ncanObj = com.open_can_channel(int(%s),int(%s))" % (self.get_CANChannelID(), self.get_CANChannelBR()))

#---------------------------- Generate Precondition Statements -------------------------------------#
    def writePreconditions(self):
        self.fo.write("\n\n")
        self.fo.write("#################################################################\n")
        self.fo.write("##      Load dbc,Periodic Tester Present,Periodic NM message   ##\n")
        self.fo.write("#################################################################\n")
        self.fo.write("\ncanObj.load_dbc(r'%s')" % self.get_DbcPath())
        self.fo.write("\ncanObj.send_cyclic_frame('BCCM_NM51F',100)")
        self.fo.write("\ncanObj.dgn.ecu_reset(0x01)" )
        self.fo.write("\ntime.sleep(1)\n" )
        self.fo.write("#################################################################\n")
#---------------------------- Generate GetActualResponseFrames() Function --------------------------#
    def writeGetActualRespDef(self):
        self.fo.write("\ndef GetActualResponseFrames():\n")
        self.fo.write("    response_str = ''\n")
        self.fo.write("    response_str = canObj.dgn.req_info_raw().split('Rsp:')[1]\n")
        self.fo.write("    if(response_str != []):\n")
        self.fo.write("        response_str = response_str.replace('[','')\n")
        self.fo.write("        response_str = response_str.replace(']','')\n")
        self.fo.write("        response_str = response_str.replace(' ','')\n")
        self.fo.write("    return response_str\n\n\n")

#---------------------------- Generate InvokeMessageBox() Function ---------------------------------#
    def writeInvokeMsgBoxDef(self):
        self.fo.write("\ndef InvokeMessageBox(MessageStr):\n")
        self.fo.write("    MessageStr = MessageStr")
        self.fo.write(r" + ")
        self.fo.write(r""" "\nPress OK after performing the action." """)
        self.fo.write("\n")
        self.fo.write("    root = tk.Tk()\n")
        self.fo.write("    root.withdraw()\n")
        self.fo.write("""    root.attributes("-topmost", True)\n""")
        self.fo.write("""    tkMessageBox.showinfo("Manual Input Required", MessageStr)\n""")
        self.fo.write("    root.destroy()\n")

#--------------------------- Python TestCase function Definition ----------------------------------#
    def writeTestCaseDef(self,testcasenumber):
            '''
            Description:This function starts to write new test case in .py file
            it writes test description ,test case requirment id to py file.
            '''
            test_name      = self.get_testname(testcasenumber)
            test_case_desc = self.get_testdescription(testcasenumber)
            test_case_reqs = self.get_testrequirementid(testcasenumber)

            self.fo.write("\n#######################")
            self.fo.write("\n## Test Case %s       ##" %(testcasenumber+1))
            self.fo.write("\n#######################"+ '\n' + '\n')
            self.fo.write('def test_%s():\n' %(str(testcasenumber+1)))

            self.TD = test_name.replace('\'','\"')  #in Test Name Statement replace ' with " for indentation
            self.TD = self.TD.replace('\n',' ')   #in Test Name Statement replace \n with ' ' for indentation
            self.TN = test_case_desc.replace('\n',' ')  #in Test Description Statement replace \n with ' ' for indentation
            self.TN = self.TN.replace('\'','\"')  #in Test Description Statement replace ' with " for indentation

            self.fo.write("    test_case_name = 'Test Case %s: %s'\n" % (testcasenumber+1,self.TD))#Write Test Name
            self.fo.write("    test_case_desc = '%s'\n" %self.TN) #Write Test Description
            self.fo.write("    test_case_reqs = '%s'\n" % test_case_reqs.replace('\n',' ') ) #Write Test Requirment ID
            self.fo.write("    print test_case_name\n\n")
            self.fo.write("    report.add_test_case(test_case_name, test_case_desc, test_case_reqs)\n\n")

#------------------------------ Process DIAG type test steps ---------------------------------------#
    def processTypeDIAG(self,testcasenumber,teststepnumber):
        #Extract values from excel sheet

        test_step_Desc    = self.get_teststepdescription(testcasenumber, teststepnumber)#Description of Test Step
        test_step_Condn   = self.get_testconditions(testcasenumber, teststepnumber)#Condition of Test Step
        test_step_ExpRes  = self.get_expectedresult(testcasenumber,teststepnumber)#Expected value of Test Step
        test_step_Comment = self.get_comments(testcasenumber, teststepnumber) #Comments of Test Step

        test_step_Desc    = test_step_Desc.replace   ('\n',' ')
        test_step_Condn   = test_step_Condn.replace  ('\n',' ')
        test_step_ExpRes  = test_step_ExpRes.replace ('\n',' ')
        test_step_Comment = test_step_Comment.replace('\n',' ')

        Expec_res  = test_step_ExpRes
        Expec_res  = Expec_res.replace(' ','')
        Expec_res  = Expec_res.upper()

        Diag_Request = test_step_Condn.replace(" ", "")

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %self.writerow_number)
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")
        self.fo.write("    ################################\n")  # Reset By Diagnostics
        self.fo.write("    #---------Service %s-----------#\n"%str(self.get_testconditions(testcasenumber,teststepnumber))[0:2])
        self.fo.write("    ################################\n\n")
        self.fo.write("\n")
        self.fo.write("\n    test_step_Desc='%s'"  %test_step_Desc)  #Test step Description
        self.fo.write("\n    test_step_Condn='%s'" %test_step_Condn)  #Test step Condition
        self.fo.write("\n    test_step_Result = False" )
        self.fo.write("\n")
        self.fo.write(r"    test_step_Comment = '\n'" )
        self.fo.write("\n")
        self.fo.write(r"    test_step_ResponseStr = '\n'" )
        self.fo.write("\n")
        self.fo.write("\n")
        self.fo.write("\n    Expec_res = '%s'\n"%Expec_res)
        self.fo.write("\n")
        self.fo.write("\n    if(Expec_res=='NONE' or Expec_res==''):")
        self.fo.write("\n        canObj.dgn.iso.net.is_spr_req = True")
        self.fo.write("\n    else:")
        self.fo.write("\n        canObj.dgn.iso.net.is_spr_req = False")
        self.fo.write("\n")

        #------------ Send Diagnostics Request bytes -----------------------
        self.fo.write("\n    datalist=[")
        for byte in range(0,len(Diag_Request),2):
            self.fo.write("%s,"%('0x'+Diag_Request[byte:byte+2]))
        self.fo.write("]")
        self.fo.write("\n")
        self.fo.write("\n    canObj.dgn.diagnostic_alltypes(datalist)")
        self.fo.write("\n    time.sleep(1)" )
        self.fo.write("\n")
        self.fo.write("\n    test_step_ActualRes = GetActualResponseFrames()")
        self.fo.write("\n")


        if(self.get_expectedresult(testcasenumber, teststepnumber).find('XX' or 'xx') >= 0):

            self.fo.write("\n    #refer to user_defined_api.py in the TEstsuite for the function def")
            self.fo.write("\n")
            self.fo.write("\n    ret = dontCareCheck_default(Expec_res, test_step_ActualRes)")
            self.fo.write("\n")
            self.fo.write("\n    if(ret==1):")
            self.fo.write("\n        test_step_Result = True")
            self.fo.write("\n        test_step_Comment ='Test Sucessful.'")
            self.fo.write("\n    else:")
            self.fo.write("\n        test_step_Result = False")
            self.fo.write("\n        test_step_Comment ='Test Fails.'")
            self.fo.write("\n")
            self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
            self.fo.write("\n")
            self.fo.write("\n")

        elif(self.get_expectedresult(testcasenumber, teststepnumber).find("SEARCH_DTC")>=0 ):

            self.fo.write("\n    #refer to user_defined_api.py in the TEstsuite for the function def")
            self.fo.write("\n")
            self.fo.write("\n    ret = searchDTC_default(Expec_res, test_step_ActualRes)")
            self.fo.write("\n")
            self.fo.write("\n    if(ret==1):")
            self.fo.write("\n        test_step_Result = True")
            self.fo.write("\n        test_step_Comment ='Test Sucessful.'")
            self.fo.write("\n    else:")
            self.fo.write("\n        test_step_Result = False")
            self.fo.write("\n        test_step_Comment ='Test Fails.'")
            self.fo.write("\n")
            self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
            self.fo.write("\n")
            self.fo.write("\n")

        elif(self.get_expectedresult(testcasenumber, teststepnumber).find("MAX")>0):

            self.fo.write("\n    #refer to user_defined_api.py in the TEstsuite for the function def")
            self.fo.write("\n")
            self.fo.write("\n    ret = rangeCheck_default(Expec_res,test_step_ActualRes)")
            self.fo.write("\n")
            self.fo.write("\n    if(ret==1):")
            self.fo.write("\n        test_step_Result = True")
            self.fo.write("\n        test_step_Comment ='Test Successful.'")
            self.fo.write("\n    else:")
            self.fo.write("\n        test_step_Result = False")
            self.fo.write("\n        test_step_Comment ='Test Fails!!'")
            self.fo.write("\n")
            self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
            self.fo.write("\n")
            self.fo.write("\n")

        elif(self.get_expectedresult(testcasenumber, teststepnumber).find("FUNC")>0 ):

            funct_name=self.get_expectedresult(testcasenumber, teststepnumber)
            funct_name = funct_name.split(':')[1]
            funct_name = funct_name.replace('"','')

            self.fo.write("\n    #refer to user_defined_api.py in the TEstsuite for the function def")
            self.fo.write("\n")
            self.fo.write("\n    ret = %s(test_step_ActualRes)"%funct_name)
            self.fo.write("\n")
            self.fo.write("\n    if(ret==1):")
            self.fo.write("\n        test_step_Result = True")
            self.fo.write("\n        test_step_Comment ='Test Sucessful.'")
            self.fo.write("\n    else:")
            self.fo.write("\n        test_step_Result = False")
            self.fo.write("\n        test_step_Comment ='Test Fails.'")
            self.fo.write("\n")
            self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
            self.fo.write("\n")
            self.fo.write("\n")

        else:
            self.fo.write("\n")
            self.fo.write("\n    if((Expec_res == test_step_ActualRes[0:len(Expec_res)]) or (Expec_res=='NONE' and len(test_step_ActualRes)<2)):" )
            self.fo.write("\n          test_step_Result = True")
            self.fo.write("\n          test_step_Comment ='Test Successful.'")
            self.fo.write("\n    else:" )
            self.fo.write("\n          test_step_Result = False")
            self.fo.write("\n          test_step_Comment ='Test Fails!!'")
            self.fo.write("\n")
            self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
            self.fo.write("\n")
            self.fo.write("\n")


#------------------------------ Process POPUP type test steps ---------------------------------------#
    def processTypePOPUP(self,testcasenumber, teststepnumber):

        test_step_Desc = self.get_teststepdescription(testcasenumber, teststepnumber) #Description of Test Step

        test_step_Desc = test_step_Desc.replace('\n',' ')
        test_step_Desc = test_step_Desc.replace('\'','\"')

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %str(self.writerow_number))
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")

        if(test_step_Desc == ''): # Check For Proper Data Entry in Excel sheet
            print("\n**Error occured.")
            print("\nTest Description is not provided.")
            print("\nCheck Row no:")+str(self.writerow_number)+"\n"
            print("\nCheck column number:")+str(self.testDescCol)+"\n"
            sys.exit()
        else:
            self.fo.write("    ################################\n")
            self.fo.write("    ##          Message Pop-up  #\n")
            self.fo.write("    ################################\n")
            self.fo.write("\n")
            self.fo.write("\n    test_step_Desc='%s'" %test_step_Desc)  #Test step Description
            self.fo.write("\n    print test_step_Desc")
            self.fo.write("\n")
            self.fo.write("\n    InvokeMessageBox(test_step_Desc)" )

        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,'','','',True,'')")
        self.fo.write("\n")
        self.fo.write("\n")

#------------------------------ Process DELAY type test steps ---------------------------------------#
    def processTypeDELAY(self,testcasenumber, teststepnumber):

        test_step_Desc    = self.get_teststepdescription(testcasenumber, teststepnumber) #Description of Test Step
        test_step_Condn   = self.get_testconditions(testcasenumber, teststepnumber) #Condition for Test Step

        test_step_Desc = test_step_Desc.replace('\n',' ') #Remove if new line is added in test test_step_Desc
        test_step_Desc = test_step_Desc.replace('\'','\"')

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %str(self.writerow_number))
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")

        if(test_step_Desc ==''): # Check For Proper Data Entry in Excel sheet
            print "\n**Error occured."
            print "\nTest Description is not provided."
            print "\nCheck Row no:" + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testDescCol)
            sys.exit()

        if(test_step_Condn == ''): #Time should not be blank
            print "\n**Error occured."
            print "\nProvide time(in sec) for delay"
            print "\nCheck Row no: " + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testCondCol)
            sys.exit()

        self.fo.write("    ################################\n")
        self.fo.write("    ##          DELAY  #\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")
        self.fo.write("\n    test_step_Desc='%s'" %test_step_Desc)
        self.fo.write("\n    test_step_Condn='%s'" %test_step_Condn)  #Test step Condition
        self.fo.write("\n    print test_step_Desc")
        self.fo.write("\n")
        self.fo.write("\n    time.sleep(%s)"%test_step_Condn)
        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'','',True,'')")
        self.fo.write("\n")
        self.fo.write("\n")


#---------------------------- Process PERIODIC_TP Test Steps -------------------------------------#
    def processTypePERIODIC_TP(self,testcasenumber,teststepnumber):

        test_step_Desc    = self.get_teststepdescription(testcasenumber, teststepnumber) #Description of Test Step
        test_step_Condn   = self.get_testconditions(testcasenumber, teststepnumber) #Condition for Test Step

        test_step_Desc = test_step_Desc.replace('\n',' ') #Remove if new line is added in test test_step_Desc
        test_step_Desc = test_step_Desc.replace('\'','\"')

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %str(self.writerow_number))
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")

        if(test_step_Desc ==''): # Check For Proper Data Entry in Excel sheet
            print "\n**Error occured."
            print "\nTest Description is not provided."
            print "\nCheck Row no:" + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testDescCol)
            sys.exit()

        if(test_step_Condn == ''): #Time should not be blank
            print "\n**Error occured."
            print "\nProvide action ON or OFF for periodic tester present"
            print "\nCheck Row no: " + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testCondCol)
            sys.exit()


        self.fo.write("    ################################\n")  # Reset By Diagnostics
        self.fo.write("    ##         PERIODIC_TP        ##\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")
        self.fo.write("\n    test_step_Desc='%s'" %test_step_Desc)
        self.fo.write("\n    test_step_Condn='%s'" %test_step_Condn)  #Test step Condition
        self.fo.write("\n    print test_step_Desc")
        self.fo.write("\n")

        if(test_step_Condn   == 'ON'):
            self.fo.write("\n    canObj.dgn.start_periodic_tp()" )
        elif(test_step_Condn == 'OFF'):
            self.fo.write("\n    canObj.dgn.stop_periodic_tp()\n")
        else:
            print "\n Invalid Value for Tester Present found !!\n"

        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'','',True,'')")
        self.fo.write("\n")
        self.fo.write("\n")


#------------------------------ Process CANSTART type test steps ---------------------------------------#
    def processTypeCANSTART(self,testcasenumber,teststepnumber):

        test_step_Desc = self.get_teststepdescription(testcasenumber, teststepnumber) #Description of Test Step

        test_step_Desc = test_step_Desc.replace('\n',' ')
        test_step_Desc = test_step_Desc.replace('\'','\"')

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s #\n" %self.writerow_number)
        self.fo.write("    #######################################\n")
        self.fo.write("\n")
        self.fo.write("    ################################\n")  # Reset By Diagnostics
        self.fo.write("    ## CAN_START COMMUNICATION      ##\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")
        self.fo.write("\n    test_step_Desc='%s'" %test_step_Desc)  #Test step Description
        self.fo.write("\n    print test_step_Desc")
        self.fo.write("\n")
        self.fo.write("\n    canObj.start_frame_reception()" )
        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,'','','',True,'')")
        self.fo.write("\n")
        self.fo.write("\n")


#------------------------------ Process CANSTOP type test steps ---------------------------------------#
    def processTypeCANSTOP(self,testcasenumber,teststepnumber):

        test_step_Desc = self.get_teststepdescription(testcasenumber, teststepnumber) #Description of Test Step

        test_step_Desc = test_step_Desc.replace('\n',' ')
        test_step_Desc = test_step_Desc.replace('\'','\"')

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s #\n" %self.writerow_number)
        self.fo.write("    #######################################\n")
        self.fo.write("\n")
        self.fo.write("    ################################\n")  # Reset By Diagnostics
        self.fo.write("    ## CAN_STOP COMMUNICATION      ##\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")
        self.fo.write("\n    test_step_Desc='%s'" %test_step_Desc)  #Test step Description
        self.fo.write("\n    print test_step_Desc")
        self.fo.write("\n")
        self.fo.write("\n    canObj.stop_frame_reception()" )
        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,'','','',True,'')")
        self.fo.write("\n")
        self.fo.write("\n")

#------------------------------ Process READ_CAN------------ ---------------------------------------#
    def processTypeRX_CAN(self,testcasenumber,teststepnumber):

        test_step_Desc    = self.get_teststepdescription(testcasenumber, teststepnumber)#Description of Test Step
        test_step_Condn   = self.get_testconditions(testcasenumber, teststepnumber)#Condition of Test Step
        test_step_ExpRes  = self.get_expectedresult(testcasenumber,teststepnumber)#Expected value of Test Step
        test_step_Comment = self.get_comments(testcasenumber, teststepnumber) #Comments of Test Step

        test_step_Desc    = test_step_Desc.replace   ('\n',' ')
        test_step_Condn   = test_step_Condn.replace  ('\n',' ')
        test_step_ExpRes  = test_step_ExpRes.replace ('\n',' ')
        test_step_Comment = test_step_Comment.replace('\n',' ')

        Expec_res  = test_step_ExpRes
        Expec_res  = Expec_res.replace(' ','')
        Expec_res  = Expec_res.upper()


        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %self.writerow_number)
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")
        self.fo.write("    ################################\n")
        self.fo.write("    #         READ CAN             #\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")

        self.fo.write("\n    test_step_Desc='%s'"  %test_step_Desc)  #Test step Description
        self.fo.write("\n    test_step_Condn='%s'" %test_step_Condn)  #Test step Condition
        self.fo.write("\n    test_step_Result = False" )
        self.fo.write("\n")
        self.fo.write(r"    test_step_Comment = '\n'" )
        self.fo.write("\n")

        if(test_step_Condn.find('.')>0):
            self.fo.write("\n")
            self.fo.write("\n    frame_name ='%s'"%str(test_step_Condn.split('.')[0]))
            self.fo.write("\n    signal_name='%s'"%str(test_step_Condn.split('.')[1].replace(" ",'')))
            self.fo.write("\n    print 'Reading %s-->>%s ' %(frame_name,signal_name)")
            self.fo.write("\n")
            self.fo.write("\n    test_step_ActualRes = canObj.get_signal(signal_name,frame_name)")
            self.fo.write("\n")

        elif(test_step_Condn != ''):
            self.fo.write("\n")
            self.fo.write("\n    frame_name ='%s'"%str(test_step_Condn))
            self.fo.write("\n    print 'Reading %s ' %frame_name")
            self.fo.write("\n")
            self.fo.write("\n    test_step_ActualRes = str(canObj.get_frame(canObj.find_frame_id(frame_name)[0]))")
            self.fo.write("\n")

        else:
            print "\n**Error occured."
            print "\nInvalid entry in Test Condition Column"
            print "\nCheck Row no: " + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testCondCol)
            sys.exit()


        self.fo.write("\n    Expec_res = '%s'\n"%Expec_res)

        self.fo.write("\n    if(Expec_res == str(test_step_ActualRes)):")
        self.fo.write("\n        test_step_Result = True")
        self.fo.write("\n        test_step_Comment ='Test Successful.'")
        self.fo.write("\n    else:")
        self.fo.write("\n        test_step_Result = False")
        self.fo.write("\n        test_step_Comment ='Test Fails.'")
        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s',test_step_ActualRes,test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
        self.fo.write("\n")
        self.fo.write("\n")

#------------------------------ Process TX_CAN------------ ---------------------------------------#
    def processTypeTX_CAN(self,testcasenumber,teststepnumber):

        test_step_Desc    = self.get_teststepdescription(testcasenumber, teststepnumber)#Description of Test Step
        test_step_Condn   = self.get_testconditions(testcasenumber, teststepnumber)#Condition of Test Step
        test_step_ExpRes  = self.get_expectedresult(testcasenumber,teststepnumber)#Expected value of Test Step
        test_step_Comment = self.get_comments(testcasenumber, teststepnumber) #Comments of Test Step

        test_step_Desc    = test_step_Desc.replace   ('\n',' ')
        test_step_Condn   = test_step_Condn.replace  ('\n',' ')
        test_step_ExpRes  = test_step_ExpRes.replace ('\n',' ')
        test_step_Comment = test_step_Comment.replace('\n',' ')

        Expec_res  = test_step_ExpRes
        Expec_res  = Expec_res.replace(' ','')
        Expec_res  = Expec_res.upper()

        self.fo.write("\n")
        self.fo.write("    #######################################\n")  #Print row number of excel sheet
        self.fo.write("    #Code block generated for row no = %s  #\n" %self.writerow_number)
        self.fo.write("    ##################################### #\n")
        self.fo.write("\n")
        self.fo.write("    ################################\n")
        self.fo.write("    #          TX CAN              #\n")
        self.fo.write("    ################################\n")
        self.fo.write("\n")

        self.fo.write("\n    test_step_Desc='%s'"  %test_step_Desc)  #Test step Description
        self.fo.write("\n    test_step_Condn='%s'" %test_step_Condn)  #Test step Condition
        self.fo.write("\n    test_step_Result = False" )
        self.fo.write("\n")
        self.fo.write(r"    test_step_Comment = '\n'" )
        self.fo.write("\n")

        if(test_step_Condn.find('.')>0):
            self.fo.write("\n")
            self.fo.write("\n    frame_name ='%s'"%str(self.get_testconditions(testcasenumber, teststepnumber).split('.')[0]))
            self.fo.write("\n    signal_name='%s'"%str(self.get_testconditions(testcasenumber, teststepnumber).split('.')[1].replace(" ",'')))
            self.fo.write("\n    print 'Writing %s-->>%s ' %(frame_name,signal_name)")
            self.fo.write("\n")
            self.fo.write("\n    ret = canObj.set_signal(signal_name,int(%s),frame_name)"%self.get_expectedresult(testcasenumber, teststepnumber))
            self.fo.write("\n    canObj.send_cyclic_frame(frame_name,10)")
        else:
            print "\n**Error occured."
            print "\nInvalid entry in Test Condition Column"
            print "\nCheck Row no: " + str(self.writerow_number)
            print "\nCheck Column no:" + str(self.testCondCol)
            sys.exit()

        self.fo.write("\n    if(ret == 1):")
        self.fo.write("\n        test_step_Result = True")
        self.fo.write("\n        test_step_Comment ='Test Successful.'")
        self.fo.write("\n    else:")
        self.fo.write("\n        test_step_Result = False")
        self.fo.write("\n        test_step_Comment ='Test Fails.'")
        self.fo.write("\n")
        self.fo.write(r"    report.add_test_step(test_step_Desc,test_step_Condn,'%s','',test_step_Result,test_step_Comment + '\n' + '%s')"%(test_step_ExpRes,test_step_Comment))
        self.fo.write("\n")
        self.fo.write("\n")

#--------------------------------------------------------------------------------------------------------------#

#-------------------------------------------- Generate EndTest Function ---------------------------------------#
    def writeEndTestDef(self):
        self.fo.write("\n\ndef endTest():")
        self.fo.write("\n    canObj.dgn.iso.net.log_file = report.get_log_dir()")
        self.fo.write("\n    canObj.dgn.save_logfile()")
        self.fo.write("\n    canObj.dgn.stop_periodic_tp()")
        self.fo.write("\n    canObj.stop_cyclic_frame('BCCM_NM51F')")
        self.fo.write("\n    report.generate_report()")
        self.fo.write("\n")
        self.fo.write("\n")
        self.fo.write(r"    print '\nScript Execution Finished !!'")
        self.fo.write("\n")
        self.fo.write("\n    com.exit()\n\n\n\n")

#------------------------------------------------ Generate Calling of Test Functions ------------------------------------------#
    def writeCallingOfTestcases(self):
        self.fo.write("\n#############################")
        self.fo.write("\n## Execute the test cases  ##")
        self.fo.write("\n#############################"+ '\n' + '\n')

        #Start Calling all Test Case
        Cnt=0
        while Cnt<self.get_totaltests():
            self.fo.write("test_%s()\n" %(Cnt+1))
            Cnt+=1
            self.fo.write("time.sleep(1)\n")

#-------------------------------------------- Finish writing the Python file -------------------------------------------------#
    def closePythonFile(self):
        self.fo.write("\n# Perform End actions i.e Save log files, Generate Report etc..")
        self.fo.write("\nendTest()")
        self.fo.close()

#------------------------------------------------------------------------------------------------------------------------------------





